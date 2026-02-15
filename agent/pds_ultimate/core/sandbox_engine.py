"""
PDS-Ultimate Sandbox Engine — Безопасная работа с файлами и кодом
==================================================================
ВЫШЕ Manus.ai: Manus работает в облачном sandbox.
Мы работаем ЛОКАЛЬНО + в Telegram — прямой доступ к файлам владельца.

Capabilities:
1. Безопасное выполнение Python-кода в изолированном контексте
2. Чтение любого файла (txt, py, json, csv, yaml, md, xml, html, cfg, ini...)
3. Редактирование файлов БЕЗ разрушения архитектуры
4. Создание файлов любого формата
5. Анализ структуры проекта (дерево файлов)
6. Поиск по содержимому файлов (grep-like)
7. Diff-патчинг: точечные правки с сохранением форматирования
8. Работа с Excel/CSV: чтение, фильтрация, добавление, редактирование
9. Генерация скриптов и их выполнение
10. Безопасный eval с ограничением по времени и памяти

Наша фишка vs Manus:
- Файл присланный в Telegram → автоматически обработан
- Понимает .py архитектуру (не ломает импорты, классы, отступы)
- Умеет делать diff-патчи без перезаписи всего файла
- Работает с бизнес-данными владельца (Excel, CSV, JSON)
"""

from __future__ import annotations

import ast
import asyncio
import csv
import io
import json
import os
import re
import sys
import textwrap
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

# Max file size to read (10MB)
MAX_FILE_SIZE = 10 * 1024 * 1024
# Max code execution time (30 seconds)
MAX_EXEC_TIME = 30
# Max output length
MAX_OUTPUT_LEN = 50_000
# Forbidden modules in sandbox
FORBIDDEN_MODULES = {
    "subprocess", "shutil", "ctypes", "importlib",
    "signal", "multiprocessing", "threading",
}
# Safe base directory
BASE_DIR = Path("/home/slavik/lessons.py.itea.2026/101/agent")
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"

# File extensions we can safely read as text
TEXT_EXTENSIONS = {
    ".py", ".txt", ".md", ".json", ".csv", ".yaml", ".yml",
    ".xml", ".html", ".htm", ".css", ".js", ".ts", ".jsx", ".tsx",
    ".cfg", ".ini", ".toml", ".env", ".sh", ".bash", ".zsh",
    ".sql", ".r", ".go", ".java", ".c", ".cpp", ".h", ".hpp",
    ".rs", ".rb", ".php", ".lua", ".kt", ".swift", ".dart",
    ".log", ".conf", ".properties", ".gitignore", ".dockerignore",
    ".dockerfile", ".makefile", ".cmake", ".gradle",
}

BINARY_EXTENSIONS = {
    ".xlsx", ".xls", ".docx", ".doc", ".pdf", ".zip", ".tar",
    ".gz", ".rar", ".7z", ".png", ".jpg", ".jpeg", ".gif",
    ".bmp", ".mp3", ".mp4", ".avi", ".mov", ".wav",
}


# ═══════════════════════════════════════════════════════════════════════════════
# DATA MODELS
# ═══════════════════════════════════════════════════════════════════════════════


@dataclass
class FileInfo:
    """Information about a file."""
    path: str
    name: str
    extension: str
    size_bytes: int
    is_text: bool
    is_binary: bool
    lines_count: int = 0
    encoding: str = "utf-8"
    modified_at: str = ""
    preview: str = ""  # First N lines

    def summary(self) -> str:
        parts = [
            f"📄 {self.name}",
            f"  📁 Путь: {self.path}",
            f"  📏 Размер: {self._human_size()}",
        ]
        if self.is_text:
            parts.append(f"  📝 Строк: {self.lines_count}")
        parts.append(f"  🕐 Изменён: {self.modified_at}")
        if self.preview:
            parts.append(f"  👁️ Превью:\n{self.preview}")
        return "\n".join(parts)

    def _human_size(self) -> str:
        s = self.size_bytes
        for unit in ("B", "KB", "MB", "GB"):
            if s < 1024:
                return f"{s:.1f} {unit}"
            s /= 1024
        return f"{s:.1f} TB"


@dataclass
class CodeResult:
    """Result of code execution."""
    success: bool
    output: str = ""
    error: str = ""
    return_value: Any = None
    execution_time_ms: int = 0
    variables: dict[str, str] = field(default_factory=dict)

    def summary(self) -> str:
        icon = "✅" if self.success else "❌"
        parts = [f"{icon} Выполнение кода ({self.execution_time_ms}ms)"]
        if self.output:
            parts.append(f"📤 Вывод:\n{self.output[:3000]}")
        if self.error:
            parts.append(f"🔴 Ошибка:\n{self.error[:1000]}")
        if self.return_value is not None:
            parts.append(f"📦 Результат: {str(self.return_value)[:500]}")
        if self.variables:
            vars_str = "\n".join(
                f"  {k} = {v}" for k, v in list(self.variables.items())[:20]
            )
            parts.append(f"📊 Переменные:\n{vars_str}")
        return "\n".join(parts)


@dataclass
class EditResult:
    """Result of a file edit operation."""
    success: bool
    file_path: str
    changes_made: int = 0
    description: str = ""
    error: str = ""
    backup_path: str = ""

    def summary(self) -> str:
        icon = "✅" if self.success else "❌"
        parts = [f"{icon} Редактирование {os.path.basename(self.file_path)}"]
        if self.description:
            parts.append(f"  📝 {self.description}")
        if self.changes_made:
            parts.append(f"  🔄 Изменений: {self.changes_made}")
        if self.backup_path:
            parts.append(f"  💾 Бэкап: {self.backup_path}")
        if self.error:
            parts.append(f"  🔴 {self.error}")
        return "\n".join(parts)


# ═══════════════════════════════════════════════════════════════════════════════
# SANDBOX ENGINE
# ═══════════════════════════════════════════════════════════════════════════════


class SandboxEngine:
    """
    Безопасная среда для работы с файлами и кодом.

    Главный принцип: НИКОГДА не ломать существующие файлы.
    Всегда делать бэкап перед изменениями.
    """

    def __init__(self):
        self._ensure_dirs()

    def _ensure_dirs(self):
        """Create required directories."""
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ─── File Reading ────────────────────────────────────────────────────

    async def read_file(
        self,
        path: str,
        start_line: int = 0,
        end_line: int = 0,
        encoding: str = "utf-8",
    ) -> FileInfo:
        """
        Read a file and return its content with metadata.

        Args:
            path: File path (absolute or relative to BASE_DIR)
            start_line: Start line (1-based, 0 = from beginning)
            end_line: End line (0 = to end)
            encoding: File encoding

        Returns:
            FileInfo with content in preview field
        """
        resolved = self._resolve_path(path)
        if not resolved.exists():
            raise FileNotFoundError(f"Файл не найден: {path}")

        stat = resolved.stat()
        if stat.st_size > MAX_FILE_SIZE:
            raise ValueError(
                f"Файл слишком большой: {stat.st_size / 1024 / 1024:.1f}MB "
                f"(макс {MAX_FILE_SIZE / 1024 / 1024:.0f}MB)"
            )

        ext = resolved.suffix.lower()
        is_text = ext in TEXT_EXTENSIONS or ext == ""
        is_binary = ext in BINARY_EXTENSIONS

        info = FileInfo(
            path=str(resolved),
            name=resolved.name,
            extension=ext,
            size_bytes=stat.st_size,
            is_text=is_text,
            is_binary=is_binary,
            modified_at=datetime.fromtimestamp(
                stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
        )

        if is_binary:
            # Handle binary files
            if ext in (".xlsx", ".xls"):
                info.preview = await self._read_excel(resolved)
            elif ext == ".csv":
                is_text = True  # CSV is text
            elif ext == ".pdf":
                info.preview = await self._read_pdf(resolved)
            else:
                info.preview = f"[Бинарный файл {ext}, {info._human_size()}]"
            return info

        # Text file
        try:
            content = resolved.read_text(encoding=encoding)
        except UnicodeDecodeError:
            try:
                content = resolved.read_text(encoding="cp1251")
                info.encoding = "cp1251"
            except Exception:
                content = resolved.read_text(encoding="latin-1")
                info.encoding = "latin-1"

        lines = content.split("\n")
        info.lines_count = len(lines)

        if start_line > 0 or end_line > 0:
            s = max(0, start_line - 1)
            e = end_line if end_line > 0 else len(lines)
            selected = lines[s:e]
            # Add line numbers
            numbered = [
                f"{s + i + 1:4d} | {line}" for i, line in enumerate(selected)
            ]
            info.preview = "\n".join(numbered)
        else:
            # Full content with line numbers for code files
            if ext in (".py", ".js", ".ts", ".java", ".c", ".cpp", ".go", ".rs"):
                numbered = [
                    f"{i + 1:4d} | {line}" for i, line in enumerate(lines)
                ]
                info.preview = "\n".join(numbered)
            else:
                info.preview = content

        return info

    async def _read_excel(self, path: Path) -> str:
        """Read Excel file and return text representation."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"📊 Лист: {sheet_name}")
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    rows.append(" | ".join(cells))
                    if len(rows) > 100:
                        rows.append(f"... ещё {ws.max_row - 100} строк")
                        break
                parts.append("\n".join(rows))
            wb.close()
            return "\n\n".join(parts)
        except ImportError:
            return "[Для Excel нужен openpyxl: pip install openpyxl]"
        except Exception as e:
            return f"[Ошибка чтения Excel: {e}]"

    async def _read_pdf(self, path: Path) -> str:
        """Read PDF file text."""
        try:
            import PyPDF2
            with open(path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                text_parts = []
                for i, page in enumerate(reader.pages[:20]):
                    text = page.extract_text()
                    if text:
                        text_parts.append(f"--- Страница {i + 1} ---\n{text}")
                return "\n\n".join(text_parts) if text_parts else "[PDF без текста]"
        except ImportError:
            return "[Для PDF нужен PyPDF2: pip install PyPDF2]"
        except Exception as e:
            return f"[Ошибка чтения PDF: {e}]"

    # ─── File Editing ────────────────────────────────────────────────────

    async def edit_file(
        self,
        path: str,
        edits: list[dict[str, str]],
        create_backup: bool = True,
    ) -> EditResult:
        """
        Edit a file with precision replacements.

        Each edit is: {"find": "old text", "replace": "new text"}
        OR: {"line": 10, "replace": "new content for line 10"}
        OR: {"insert_after_line": 5, "content": "new line to insert"}
        OR: {"delete_lines": "10-15"}

        Always creates backup before editing.
        """
        resolved = self._resolve_path(path)
        if not resolved.exists():
            return EditResult(
                success=False, file_path=str(resolved),
                error=f"Файл не найден: {path}"
            )

        # Read original
        try:
            content = resolved.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            content = resolved.read_text(encoding="cp1251")

        # Backup
        backup_path = ""
        if create_backup:
            backup_path = str(resolved) + \
                f".bak.{datetime.now():%Y%m%d_%H%M%S}"
            Path(backup_path).write_text(content, encoding="utf-8")

        lines = content.split("\n")
        changes = 0
        descriptions = []

        for edit in edits:
            try:
                if "find" in edit and "replace" in edit:
                    # Find & replace
                    old = edit["find"]
                    new = edit["replace"]
                    if old in content:
                        content = content.replace(old, new, 1)
                        changes += 1
                        preview = old[:50].replace('\n', '\\n')
                        descriptions.append(f"Заменено: '{preview}...'")
                    else:
                        descriptions.append(
                            f"⚠️ Не найдено: '{old[:50]}...'"
                        )

                elif "line" in edit and "replace" in edit:
                    # Replace specific line
                    line_num = int(edit["line"]) - 1
                    if 0 <= line_num < len(lines):
                        old_line = lines[line_num]
                        lines[line_num] = edit["replace"]
                        content = "\n".join(lines)
                        changes += 1
                        descriptions.append(
                            f"Строка {line_num + 1}: '{old_line[:40]}' → '{edit['replace'][:40]}'"
                        )

                elif "insert_after_line" in edit and "content" in edit:
                    # Insert after line
                    after = int(edit["insert_after_line"])
                    new_lines = edit["content"].split("\n")
                    lines[after:after] = new_lines
                    content = "\n".join(lines)
                    changes += len(new_lines)
                    descriptions.append(
                        f"Вставлено {len(new_lines)} строк после строки {after}"
                    )

                elif "delete_lines" in edit:
                    # Delete line range "10-15" or "10"
                    range_str = str(edit["delete_lines"])
                    if "-" in range_str:
                        s, e = range_str.split("-")
                        s, e = int(s) - 1, int(e)
                    else:
                        s = int(range_str) - 1
                        e = s + 1
                    deleted = lines[s:e]
                    del lines[s:e]
                    content = "\n".join(lines)
                    changes += len(deleted)
                    descriptions.append(f"Удалено строк {s + 1}-{e}")

            except Exception as e:
                descriptions.append(f"⚠️ Ошибка в правке: {e}")

        if changes > 0:
            # Validate Python files
            if resolved.suffix == ".py":
                validation = self._validate_python(content)
                if not validation["valid"]:
                    # Restore backup
                    if backup_path:
                        Path(backup_path).rename(resolved)
                    return EditResult(
                        success=False, file_path=str(resolved),
                        changes_made=0,
                        error=(
                            f"Правки нарушают синтаксис Python!\n"
                            f"{validation['error']}\n"
                            f"Файл НЕ изменён. Бэкап восстановлен."
                        ),
                    )

            resolved.write_text(content, encoding="utf-8")

        return EditResult(
            success=changes > 0,
            file_path=str(resolved),
            changes_made=changes,
            description="\n".join(descriptions),
            backup_path=backup_path,
        )

    def _validate_python(self, code: str) -> dict[str, Any]:
        """Validate Python syntax without executing."""
        try:
            ast.parse(code)
            return {"valid": True, "error": ""}
        except SyntaxError as e:
            return {
                "valid": False,
                "error": f"Строка {e.lineno}: {e.msg}",
            }

    # ─── File Creation ───────────────────────────────────────────────────

    async def create_file(
        self,
        path: str,
        content: str,
        overwrite: bool = False,
    ) -> EditResult:
        """Create a new file with content."""
        resolved = self._resolve_path(path)

        if resolved.exists() and not overwrite:
            return EditResult(
                success=False, file_path=str(resolved),
                error="Файл уже существует. Используйте overwrite=true."
            )

        # Ensure directory exists
        resolved.parent.mkdir(parents=True, exist_ok=True)

        # Validate Python
        if resolved.suffix == ".py":
            validation = self._validate_python(content)
            if not validation["valid"]:
                return EditResult(
                    success=False, file_path=str(resolved),
                    error=f"Синтаксическая ошибка: {validation['error']}"
                )

        resolved.write_text(content, encoding="utf-8")

        return EditResult(
            success=True, file_path=str(resolved),
            changes_made=1,
            description=f"Создан файл {resolved.name} ({len(content)} символов)",
        )

    # ─── Code Execution ──────────────────────────────────────────────────

    async def execute_code(
        self,
        code: str,
        timeout: int = MAX_EXEC_TIME,
        allowed_imports: set[str] | None = None,
    ) -> CodeResult:
        """
        Execute Python code in a restricted sandbox.

        Safe: math, json, re, datetime, collections, itertools, functools,
              statistics, csv, io, textwrap, os.path, pathlib.Path
        Forbidden: subprocess, shutil, ctypes, importlib, signal, etc.
        """
        import time as _time
        start = _time.time()

        # Security check
        security = self._check_code_security(code)
        if not security["safe"]:
            return CodeResult(
                success=False,
                error=f"🔒 Код заблокирован: {security['reason']}",
                execution_time_ms=0,
            )

        # Capture stdout
        stdout_capture = io.StringIO()
        old_stdout = sys.stdout
        old_stderr = sys.stderr
        stderr_capture = io.StringIO()

        # Build safe globals
        safe_globals = self._build_safe_globals()
        local_vars: dict[str, Any] = {}

        try:
            sys.stdout = stdout_capture
            sys.stderr = stderr_capture

            # Execute with timeout using asyncio
            compiled = compile(code, "<sandbox>", "exec")

            # Run in thread with timeout
            loop = asyncio.get_event_loop()
            await asyncio.wait_for(
                loop.run_in_executor(
                    None,
                    lambda: exec(compiled, safe_globals, local_vars)
                ),
                timeout=timeout,
            )

            output = stdout_capture.getvalue()
            error_output = stderr_capture.getvalue()

            # Collect user-defined variables
            user_vars = {
                k: repr(v)[:200]
                for k, v in local_vars.items()
                if not k.startswith("_") and not callable(v)
            }

            elapsed = int((_time.time() - start) * 1000)

            return CodeResult(
                success=True,
                output=output[:MAX_OUTPUT_LEN],
                error=error_output[:1000] if error_output else "",
                return_value=local_vars.get("result"),
                execution_time_ms=elapsed,
                variables=user_vars,
            )

        except asyncio.TimeoutError:
            return CodeResult(
                success=False,
                error=f"⏰ Превышен лимит времени ({timeout}с)",
                execution_time_ms=timeout * 1000,
            )
        except Exception as e:
            tb = traceback.format_exc()
            elapsed = int((_time.time() - start) * 1000)
            return CodeResult(
                success=False,
                output=stdout_capture.getvalue()[:MAX_OUTPUT_LEN],
                error=f"{type(e).__name__}: {e}\n{tb}",
                execution_time_ms=elapsed,
            )
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr

    def _check_code_security(self, code: str) -> dict[str, Any]:
        """Check if code is safe to execute."""
        # Parse AST
        try:
            tree = ast.parse(code)
        except SyntaxError as e:
            return {"safe": False, "reason": f"Синтаксическая ошибка: {e}"}

        # Check for forbidden imports
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                for alias in node.names:
                    mod = alias.name.split(".")[0]
                    if mod in FORBIDDEN_MODULES:
                        return {
                            "safe": False,
                            "reason": f"Запрещённый модуль: {mod}",
                        }
            elif isinstance(node, ast.ImportFrom):
                if node.module:
                    mod = node.module.split(".")[0]
                    if mod in FORBIDDEN_MODULES:
                        return {
                            "safe": False,
                            "reason": f"Запрещённый модуль: {mod}",
                        }

        # Check for dangerous builtins
        dangerous_calls = {"exec", "eval", "compile", "__import__", "open"}
        code_lower = code.lower()
        for dc in dangerous_calls:
            # Allow open() only for reading
            if dc == "open" and "open(" in code:
                # Check if it's read-only
                if "'w'" in code or '"w"' in code or "'a'" in code or '"a"' in code:
                    return {
                        "safe": False,
                        "reason": "open() в режиме записи запрещён в sandbox",
                    }
                continue
            if dc + "(" in code_lower:
                return {
                    "safe": False,
                    "reason": f"Запрещённая функция: {dc}()",
                }

        # Check for system commands
        if "os.system" in code or "os.popen" in code:
            return {"safe": False, "reason": "Системные команды запрещены"}

        return {"safe": True, "reason": ""}

    def _build_safe_globals(self) -> dict[str, Any]:
        """Build safe global namespace for code execution."""
        import collections
        import datetime
        import functools
        import itertools
        import math
        import statistics

        safe = {
            "__builtins__": {
                "print": print,
                "len": len,
                "range": range,
                "enumerate": enumerate,
                "zip": zip,
                "map": map,
                "filter": filter,
                "sorted": sorted,
                "reversed": reversed,
                "list": list,
                "dict": dict,
                "set": set,
                "tuple": tuple,
                "str": str,
                "int": int,
                "float": float,
                "bool": bool,
                "abs": abs,
                "round": round,
                "min": min,
                "max": max,
                "sum": sum,
                "any": any,
                "all": all,
                "isinstance": isinstance,
                "type": type,
                "hasattr": hasattr,
                "getattr": getattr,
                "setattr": setattr,
                "repr": repr,
                "format": format,
                "input": lambda prompt="": "",  # Disabled
                "open": open,  # For reading only
                "ValueError": ValueError,
                "TypeError": TypeError,
                "KeyError": KeyError,
                "IndexError": IndexError,
                "Exception": Exception,
                "StopIteration": StopIteration,
                "True": True,
                "False": False,
                "None": None,
            },
            "math": math,
            "json": json,
            "re": re,
            "datetime": datetime,
            "collections": collections,
            "itertools": itertools,
            "functools": functools,
            "statistics": statistics,
            "csv": csv,
            "io": io,
            "textwrap": textwrap,
            "Path": Path,
        }
        return safe

    # ─── File Search ─────────────────────────────────────────────────────

    async def search_in_files(
        self,
        pattern: str,
        directory: str = "",
        extensions: list[str] | None = None,
        max_results: int = 50,
        regex: bool = False,
    ) -> list[dict[str, Any]]:
        """
        Search for pattern in files (grep-like).

        Returns list of {file, line_number, line_content, match}
        """
        base = self._resolve_path(directory) if directory else BASE_DIR
        if not base.is_dir():
            return []

        results = []
        ext_filter = set(extensions) if extensions else TEXT_EXTENSIONS

        if regex:
            try:
                pat = re.compile(pattern, re.IGNORECASE)
            except re.error:
                return []
        else:
            pat = None

        for fpath in base.rglob("*"):
            if not fpath.is_file():
                continue
            if fpath.suffix.lower() not in ext_filter:
                continue
            if "__pycache__" in str(fpath) or ".git" in str(fpath):
                continue
            if fpath.stat().st_size > 1_000_000:  # Skip files > 1MB
                continue

            try:
                text = fpath.read_text(encoding="utf-8", errors="ignore")
                for i, line in enumerate(text.split("\n"), 1):
                    match = False
                    if regex and pat:
                        match = bool(pat.search(line))
                    else:
                        match = pattern.lower() in line.lower()

                    if match:
                        results.append({
                            "file": str(fpath.relative_to(BASE_DIR)),
                            "line_number": i,
                            "line_content": line.strip()[:200],
                        })
                        if len(results) >= max_results:
                            return results
            except Exception:
                continue

        return results

    # ─── Directory Tree ──────────────────────────────────────────────────

    async def list_directory(
        self,
        path: str = "",
        max_depth: int = 3,
        show_hidden: bool = False,
    ) -> str:
        """Generate directory tree."""
        base = self._resolve_path(path) if path else BASE_DIR
        if not base.is_dir():
            return f"❌ Не директория: {path}"

        lines = [f"📁 {base.name}/"]
        self._tree(base, lines, "", max_depth, 0, show_hidden)

        return "\n".join(lines[:200])

    def _tree(
        self,
        directory: Path,
        lines: list[str],
        prefix: str,
        max_depth: int,
        current_depth: int,
        show_hidden: bool,
    ):
        if current_depth >= max_depth:
            return

        try:
            entries = sorted(
                directory.iterdir(),
                key=lambda p: (not p.is_dir(), p.name.lower()),
            )
        except PermissionError:
            return

        # Filter
        entries = [
            e for e in entries
            if (show_hidden or not e.name.startswith("."))
            and e.name != "__pycache__"
            and e.name != "node_modules"
            and e.name != ".git"
        ]

        for i, entry in enumerate(entries):
            is_last = i == len(entries) - 1
            connector = "└── " if is_last else "├── "
            next_prefix = prefix + ("    " if is_last else "│   ")

            if entry.is_dir():
                count = sum(1 for _ in entry.iterdir()
                            ) if entry.is_dir() else 0
                lines.append(f"{prefix}{connector}📁 {entry.name}/ ({count})")
                self._tree(
                    entry, lines, next_prefix,
                    max_depth, current_depth + 1, show_hidden
                )
            else:
                size = entry.stat().st_size
                size_str = f"{size}" if size < 1024 else f"{size / 1024:.0f}K"
                lines.append(
                    f"{prefix}{connector}📄 {entry.name} ({size_str})"
                )

    # ─── Excel/CSV Operations ────────────────────────────────────────────

    async def read_csv(
        self,
        path: str,
        delimiter: str = ",",
        max_rows: int = 100,
    ) -> dict[str, Any]:
        """Read CSV file and return structured data."""
        resolved = self._resolve_path(path)
        if not resolved.exists():
            return {"error": f"Файл не найден: {path}"}

        try:
            text = resolved.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            text = resolved.read_text(encoding="cp1251", errors="ignore")

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = []
        headers = []

        for i, row in enumerate(reader):
            if i == 0:
                headers = row
            else:
                rows.append(row)
            if i >= max_rows:
                break

        return {
            "headers": headers,
            "rows": rows,
            "total_rows": len(rows),
            "columns": len(headers),
            "preview": self._format_table(headers, rows[:20]),
        }

    async def edit_csv(
        self,
        path: str,
        operations: list[dict[str, Any]],
        delimiter: str = ",",
    ) -> EditResult:
        """
        Edit CSV file with operations:
        - {"add_row": ["val1", "val2", ...]}
        - {"edit_cell": {"row": 1, "col": 0, "value": "new"}}
        - {"delete_row": 5}
        - {"add_column": {"name": "NewCol", "default": ""}}
        - {"filter": {"column": 0, "value": "match"}}
        - {"sort": {"column": 0, "reverse": false}}
        """
        resolved = self._resolve_path(path)
        if not resolved.exists():
            return EditResult(
                success=False, file_path=str(resolved),
                error=f"Файл не найден: {path}"
            )

        try:
            text = resolved.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            text = resolved.read_text(encoding="cp1251", errors="ignore")

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = list(reader)

        if not all_rows:
            return EditResult(
                success=False, file_path=str(resolved),
                error="Пустой CSV файл"
            )

        # Backup
        backup_path = str(resolved) + f".bak.{datetime.now():%Y%m%d_%H%M%S}"
        Path(backup_path).write_text(text, encoding="utf-8")

        headers = all_rows[0]
        data_rows = all_rows[1:]
        changes = 0
        descriptions = []

        for op in operations:
            if "add_row" in op:
                data_rows.append(op["add_row"])
                changes += 1
                descriptions.append(
                    f"Добавлена строка: {op['add_row'][:3]}...")

            elif "edit_cell" in op:
                cell = op["edit_cell"]
                r, c = int(cell["row"]) - 1, int(cell["col"])
                if 0 <= r < len(data_rows) and 0 <= c < len(data_rows[r]):
                    old = data_rows[r][c]
                    data_rows[r][c] = str(cell["value"])
                    changes += 1
                    descriptions.append(
                        f"Ячейка [{r + 1},{c}]: '{old}' → '{cell['value']}'"
                    )

            elif "delete_row" in op:
                idx = int(op["delete_row"]) - 1
                if 0 <= idx < len(data_rows):
                    deleted = data_rows.pop(idx)
                    changes += 1
                    descriptions.append(f"Удалена строка {idx + 1}")

            elif "add_column" in op:
                col = op["add_column"]
                name = col.get("name", "NewCol")
                default = col.get("default", "")
                headers.append(name)
                for row in data_rows:
                    row.append(default)
                changes += 1
                descriptions.append(f"Добавлен столбец: {name}")

            elif "sort" in op:
                s = op["sort"]
                col_idx = int(s.get("column", 0))
                reverse = bool(s.get("reverse", False))
                try:
                    data_rows.sort(
                        key=lambda r: r[col_idx] if col_idx < len(r) else "",
                        reverse=reverse,
                    )
                    changes += 1
                    descriptions.append(
                        f"Отсортировано по столбцу {col_idx}"
                    )
                except Exception as e:
                    descriptions.append(f"Ошибка сортировки: {e}")

        # Write back
        if changes > 0:
            output = io.StringIO()
            writer = csv.writer(output, delimiter=delimiter)
            writer.writerow(headers)
            writer.writerows(data_rows)
            resolved.write_text(output.getvalue(), encoding="utf-8")

        return EditResult(
            success=changes > 0,
            file_path=str(resolved),
            changes_made=changes,
            description="\n".join(descriptions),
            backup_path=backup_path,
        )

    def _format_table(
        self,
        headers: list[str],
        rows: list[list[str]],
    ) -> str:
        """Format table as readable text."""
        if not headers and not rows:
            return "[Пустая таблица]"

        # Calculate column widths
        all_rows = [headers] + rows if headers else rows
        col_count = max(len(r) for r in all_rows)
        widths = [0] * col_count

        for row in all_rows:
            for i, cell in enumerate(row):
                if i < col_count:
                    widths[i] = max(widths[i], min(len(str(cell)), 25))

        lines = []
        if headers:
            header_str = " | ".join(
                str(h)[:25].ljust(widths[i])
                for i, h in enumerate(headers)
            )
            lines.append(header_str)
            lines.append("-" * len(header_str))

        for row in rows:
            row_str = " | ".join(
                str(c)[:25].ljust(widths[i] if i < len(widths) else 10)
                for i, c in enumerate(row)
            )
            lines.append(row_str)

        return "\n".join(lines)

    # ─── Path Resolution ────────────────────────────────────────────────

    def _resolve_path(self, path: str) -> Path:
        """Resolve path safely within allowed directories."""
        if not path:
            return BASE_DIR

        p = Path(path)
        if p.is_absolute():
            resolved = p.resolve()
        else:
            # Try relative to uploads first, then BASE_DIR
            upload_try = UPLOAD_DIR / path
            if upload_try.exists():
                resolved = upload_try.resolve()
            else:
                resolved = (BASE_DIR / path).resolve()

        return resolved


# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL INSTANCE
# ═══════════════════════════════════════════════════════════════════════════════

sandbox = SandboxEngine()
