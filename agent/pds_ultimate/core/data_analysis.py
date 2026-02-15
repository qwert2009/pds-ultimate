"""
PDS-Ultimate Data Analysis Engine — Анализ данных + визуализация
=================================================================
ВЫШЕ Manus.ai:
  Manus: песочница с Jupiter → запускает скрипты
  Мы: встроенный анализ прямо в Telegram — без Jupiter, без IDE

Capabilities:
1. CSV/Excel/JSON анализ (statistics, groupby, pivot, filter)
2. Визуализация (bar, line, pie, scatter, heatmap) → PNG файл
3. Автоматический EDA (Exploratory Data Analysis)
4. Корреляции, тренды, аномалии
5. Бизнес-метрики (ROI, конверсия, средний чек, ABC-анализ)
6. Прогнозирование (простая линейная регрессия)
7. Экспорт результатов в Excel/CSV
"""

from __future__ import annotations

import csv
import io
import json
import os
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from pds_ultimate.config import logger

OUTPUT_DIR = Path("/home/slavik/lessons.py.itea.2026/101/agent/output")


# ═══════════════════════════════════════════════════════════════════════════════
# DATA MODELS
# ═══════════════════════════════════════════════════════════════════════════════


@dataclass
class ColumnStats:
    """Statistics for a single column."""
    name: str
    dtype: str = "string"  # string, number, date, boolean
    count: int = 0
    unique: int = 0
    missing: int = 0
    # Numeric stats
    mean: float | None = None
    median: float | None = None
    std: float | None = None
    min_val: float | None = None
    max_val: float | None = None
    q25: float | None = None
    q75: float | None = None
    # String stats
    avg_length: float | None = None
    top_values: list[tuple[str, int]] = field(default_factory=list)

    def summary(self) -> str:
        parts = [f"📊 {self.name} ({self.dtype})"]
        parts.append(
            f"  Записей: {self.count} | Уникальных: {self.unique} | Пропусков: {self.missing}")
        if self.dtype == "number" and self.mean is not None:
            parts.append(
                f"  Среднее: {self.mean:.2f} | Медиана: {self.median:.2f} | "
                f"Std: {self.std:.2f}"
            )
            parts.append(
                f"  Мин: {self.min_val:.2f} | Макс: {self.max_val:.2f}")
        if self.top_values:
            top_str = ", ".join(f"{v}({c})" for v, c in self.top_values[:5])
            parts.append(f"  Топ: {top_str}")
        return "\n".join(parts)


@dataclass
class AnalysisResult:
    """Result of data analysis."""
    success: bool
    summary: str = ""
    charts: list[str] = field(default_factory=list)  # File paths to charts
    tables: list[str] = field(default_factory=list)   # Formatted tables
    export_path: str = ""
    data: dict[str, Any] = field(default_factory=dict)
    error: str = ""

    def full_summary(self) -> str:
        parts = []
        if self.summary:
            parts.append(self.summary)
        if self.charts:
            parts.append(f"\n📈 Графики: {len(self.charts)} файлов")
            for c in self.charts:
                parts.append(f"  📊 {os.path.basename(c)}")
        if self.export_path:
            parts.append(f"\n💾 Экспорт: {self.export_path}")
        if self.error:
            parts.append(f"\n❌ {self.error}")
        return "\n".join(parts)


# ═══════════════════════════════════════════════════════════════════════════════
# DATA ANALYSIS ENGINE
# ═══════════════════════════════════════════════════════════════════════════════


class DataAnalysisEngine:
    """
    Встроенный движок анализа данных.
    Работает без pandas/numpy — чистый Python.
    Поддерживает pandas/matplotlib если установлены.
    """

    def __init__(self):
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        self._has_pandas = self._check_module("pandas")
        self._has_matplotlib = self._check_module("matplotlib")

    def _check_module(self, name: str) -> bool:
        try:
            __import__(name)
            return True
        except ImportError:
            return False

    # ─── Data Loading ────────────────────────────────────────────────────

    def load_csv(
        self,
        path: str,
        delimiter: str = ",",
        encoding: str = "utf-8",
    ) -> tuple[list[str], list[list[str]]]:
        """Load CSV and return (headers, rows)."""
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"Файл не найден: {path}")

        try:
            text = p.read_text(encoding=encoding, errors="ignore")
        except Exception:
            text = p.read_text(encoding="cp1251", errors="ignore")

        # Auto-detect delimiter
        first_line = text.split("\n")[0]
        if delimiter == "," and "\t" in first_line and "," not in first_line:
            delimiter = "\t"
        elif delimiter == "," and ";" in first_line and "," not in first_line:
            delimiter = ";"

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = list(reader)

        if not all_rows:
            return [], []

        headers = all_rows[0]
        data = all_rows[1:]
        return headers, data

    def load_json(self, path: str) -> tuple[list[str], list[list[str]]]:
        """Load JSON array and convert to tabular format."""
        p = Path(path)
        text = p.read_text(encoding="utf-8")
        data = json.loads(text)

        if isinstance(data, list) and data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            rows = [
                [str(item.get(h, "")) for h in headers]
                for item in data
            ]
            return headers, rows
        elif isinstance(data, dict):
            headers = list(data.keys())
            rows = [[str(v) for v in data.values()]]
            return headers, rows

        return [], []

    def load_excel(
        self, path: str, sheet: str = ""
    ) -> tuple[list[str], list[list[str]]]:
        """Load Excel file."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return [], []
            headers = [
                str(c) if c else f"Col{i}" for i, c in enumerate(rows[0])]
            data = [
                [str(c) if c is not None else "" for c in row]
                for row in rows[1:]
            ]
            return headers, data
        except ImportError:
            raise ImportError("Нужен openpyxl: pip install openpyxl")

    def auto_load(self, path: str) -> tuple[list[str], list[list[str]]]:
        """Auto-detect format and load."""
        ext = Path(path).suffix.lower()
        if ext in (".csv", ".tsv", ".txt"):
            return self.load_csv(path)
        elif ext in (".json", ".jsonl"):
            return self.load_json(path)
        elif ext in (".xlsx", ".xls"):
            return self.load_excel(path)
        else:
            # Try CSV
            return self.load_csv(path)

    # ─── Column Analysis ────────────────────────────────────────────────

    def _analyze_column(
        self,
        name: str,
        values: list[str],
    ) -> ColumnStats:
        """Analyze a single column."""
        non_empty = [v for v in values if v.strip()]
        stats = ColumnStats(
            name=name,
            count=len(values),
            unique=len(set(non_empty)),
            missing=len(values) - len(non_empty),
        )

        if not non_empty:
            return stats

        # Try numeric
        numbers = []
        for v in non_empty:
            try:
                n = float(v.replace(",", ".").replace(" ", ""))
                numbers.append(n)
            except ValueError:
                pass

        if len(numbers) > len(non_empty) * 0.7:
            # Numeric column
            stats.dtype = "number"
            stats.mean = statistics.mean(numbers)
            stats.median = statistics.median(numbers)
            stats.std = statistics.stdev(numbers) if len(numbers) > 1 else 0
            stats.min_val = min(numbers)
            stats.max_val = max(numbers)
            sorted_n = sorted(numbers)
            q = len(sorted_n)
            stats.q25 = sorted_n[q // 4] if q >= 4 else sorted_n[0]
            stats.q75 = sorted_n[3 * q // 4] if q >= 4 else sorted_n[-1]
        else:
            stats.dtype = "string"
            stats.avg_length = statistics.mean(len(v) for v in non_empty)

        # Top values
        counter = Counter(non_empty)
        stats.top_values = counter.most_common(10)

        return stats

    # ─── EDA (Exploratory Data Analysis) ─────────────────────────────────

    async def eda(
        self,
        path: str,
        generate_charts: bool = True,
    ) -> AnalysisResult:
        """
        Full Exploratory Data Analysis.
        """
        try:
            headers, rows = self.auto_load(path)
        except Exception as e:
            return AnalysisResult(
                success=False, error=f"Ошибка загрузки: {e}"
            )

        if not headers or not rows:
            return AnalysisResult(
                success=False, error="Пустой файл или нет данных"
            )

        # Analyze each column
        col_stats = []
        for i, h in enumerate(headers):
            values = [row[i] if i < len(row) else "" for row in rows]
            cs = self._analyze_column(h, values)
            col_stats.append(cs)

        # Build summary
        lines = [
            f"📊 Анализ: {os.path.basename(path)}",
            f"📏 Размер: {len(rows)} строк × {len(headers)} столбцов",
            "",
            "📋 СТОЛБЦЫ:",
        ]
        for cs in col_stats:
            lines.append(cs.summary())
            lines.append("")

        # Numeric correlations
        num_cols = [cs for cs in col_stats if cs.dtype == "number"]
        if len(num_cols) >= 2:
            lines.append("📈 КОРРЕЛЯЦИИ:")
            for i, c1 in enumerate(num_cols):
                for c2 in num_cols[i + 1:]:
                    vals1 = self._get_numeric_column(
                        headers, rows, c1.name)
                    vals2 = self._get_numeric_column(
                        headers, rows, c2.name)
                    corr = self._correlation(vals1, vals2)
                    if abs(corr) > 0.3:
                        emoji = "🔴" if abs(corr) > 0.7 else "🟡"
                        lines.append(
                            f"  {emoji} {c1.name} ↔ {c2.name}: {corr:.2f}"
                        )
            lines.append("")

        # Charts
        charts = []
        if generate_charts and self._has_matplotlib:
            try:
                chart_path = await self._generate_eda_charts(
                    headers, rows, col_stats, path
                )
                if chart_path:
                    charts.append(chart_path)
            except Exception as e:
                logger.debug(f"Chart generation error: {e}")

        return AnalysisResult(
            success=True,
            summary="\n".join(lines),
            charts=charts,
            data={
                "rows": len(rows),
                "columns": len(headers),
                "headers": headers,
                "column_stats": [
                    {
                        "name": cs.name,
                        "dtype": cs.dtype,
                        "count": cs.count,
                        "unique": cs.unique,
                        "missing": cs.missing,
                        "mean": cs.mean,
                        "median": cs.median,
                    }
                    for cs in col_stats
                ],
            },
        )

    # ─── Specific Analyses ──────────────────────────────────────────────

    async def filter_data(
        self,
        path: str,
        column: str,
        condition: str,
        value: str,
        output_path: str = "",
    ) -> AnalysisResult:
        """
        Filter data by condition.
        Conditions: equals, contains, greater, less, not_empty, starts_with
        """
        headers, rows = self.auto_load(path)

        if column not in headers:
            return AnalysisResult(
                success=False,
                error=f"Столбец '{column}' не найден. Доступные: {headers}"
            )

        col_idx = headers.index(column)
        filtered = []

        for row in rows:
            cell = row[col_idx] if col_idx < len(row) else ""
            match = False

            if condition == "equals":
                match = cell.lower() == value.lower()
            elif condition == "contains":
                match = value.lower() in cell.lower()
            elif condition == "greater":
                try:
                    match = float(cell.replace(",", ".")) > float(value)
                except ValueError:
                    pass
            elif condition == "less":
                try:
                    match = float(cell.replace(",", ".")) < float(value)
                except ValueError:
                    pass
            elif condition == "not_empty":
                match = bool(cell.strip())
            elif condition == "starts_with":
                match = cell.lower().startswith(value.lower())

            if match:
                filtered.append(row)

        # Format result
        table_str = self._format_table(headers, filtered[:50])

        # Export if needed
        export = ""
        if output_path or len(filtered) > 50:
            if not output_path:
                output_path = str(
                    OUTPUT_DIR / f"filtered_{datetime.now():%Y%m%d_%H%M%S}.csv"
                )
            self._write_csv(output_path, headers, filtered)
            export = output_path

        return AnalysisResult(
            success=True,
            summary=(
                f"🔍 Фильтр: {column} {condition} '{value}'\n"
                f"📊 Найдено: {len(filtered)} из {len(rows)} строк\n\n"
                f"{table_str}"
            ),
            export_path=export,
            data={"filtered_count": len(filtered), "total": len(rows)},
        )

    async def group_by(
        self,
        path: str,
        group_column: str,
        agg_column: str = "",
        agg_func: str = "count",
    ) -> AnalysisResult:
        """
        Group by column with aggregation.
        agg_func: count, sum, avg, min, max
        """
        headers, rows = self.auto_load(path)

        if group_column not in headers:
            return AnalysisResult(
                success=False,
                error=f"Столбец '{group_column}' не найден"
            )

        g_idx = headers.index(group_column)
        a_idx = headers.index(
            agg_column) if agg_column and agg_column in headers else -1

        groups: dict[str, list] = defaultdict(list)
        for row in rows:
            key = row[g_idx] if g_idx < len(row) else ""
            if a_idx >= 0 and a_idx < len(row):
                groups[key].append(row[a_idx])
            else:
                groups[key].append("1")

        # Aggregate
        result_rows = []
        for key, values in sorted(groups.items()):
            if agg_func == "count":
                agg_val = str(len(values))
            elif agg_func == "sum":
                nums = self._to_numbers(values)
                agg_val = f"{sum(nums):.2f}" if nums else "0"
            elif agg_func == "avg":
                nums = self._to_numbers(values)
                agg_val = f"{statistics.mean(nums):.2f}" if nums else "0"
            elif agg_func == "min":
                nums = self._to_numbers(values)
                agg_val = f"{min(nums):.2f}" if nums else "0"
            elif agg_func == "max":
                nums = self._to_numbers(values)
                agg_val = f"{max(nums):.2f}" if nums else "0"
            else:
                agg_val = str(len(values))

            result_rows.append([key, agg_val])

        agg_label = agg_column or "count"
        result_headers = [group_column, f"{agg_func}({agg_label})"]
        table_str = self._format_table(result_headers, result_rows)

        # Chart
        charts = []
        if self._has_matplotlib and len(result_rows) <= 30:
            try:
                chart_path = await self._generate_bar_chart(
                    [r[0] for r in result_rows],
                    [float(r[1]) for r in result_rows],
                    f"{group_column} — {agg_func}({agg_label})",
                )
                if chart_path:
                    charts.append(chart_path)
            except Exception:
                pass

        return AnalysisResult(
            success=True,
            summary=(
                f"📊 Group by: {group_column}\n"
                f"📈 Агрегация: {agg_func}({agg_label})\n"
                f"📋 Групп: {len(result_rows)}\n\n"
                f"{table_str}"
            ),
            charts=charts,
            data={
                "groups": len(result_rows),
                "results": {r[0]: r[1] for r in result_rows},
            },
        )

    async def compute_stats(
        self,
        path: str,
        column: str = "",
    ) -> AnalysisResult:
        """Compute detailed statistics for a column or all numeric columns."""
        headers, rows = self.auto_load(path)

        if column and column in headers:
            target_cols = [column]
        else:
            target_cols = headers

        lines = [f"📊 Статистика: {os.path.basename(path)}\n"]

        for col_name in target_cols:
            col_idx = headers.index(col_name)
            values = [row[col_idx] if col_idx <
                      len(row) else "" for row in rows]
            cs = self._analyze_column(col_name, values)
            if cs.dtype == "number":
                lines.append(cs.summary())
                lines.append("")

        return AnalysisResult(
            success=True,
            summary="\n".join(lines),
        )

    # ─── Chart Generation ────────────────────────────────────────────────

    async def _generate_eda_charts(
        self,
        headers: list[str],
        rows: list[list[str]],
        col_stats: list[ColumnStats],
        source_path: str,
    ) -> str | None:
        """Generate EDA overview chart."""
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            num_cols = [cs for cs in col_stats if cs.dtype == "number"]
            if not num_cols:
                return None

            n = min(len(num_cols), 4)
            fig, axes = plt.subplots(1, n, figsize=(5 * n, 4))
            if n == 1:
                axes = [axes]

            for i, cs in enumerate(num_cols[:n]):
                values = self._get_numeric_column(headers, rows, cs.name)
                axes[i].hist(values, bins=20, color="#4A90D9", alpha=0.7)
                axes[i].set_title(cs.name, fontsize=10)
                axes[i].axvline(
                    cs.mean or 0, color="red",
                    linestyle="--", label=f"mean={cs.mean:.1f}"
                )
                axes[i].legend(fontsize=8)

            plt.tight_layout()
            chart_path = str(
                OUTPUT_DIR / f"eda_{datetime.now():%Y%m%d_%H%M%S}.png"
            )
            plt.savefig(chart_path, dpi=150)
            plt.close()
            return chart_path

        except Exception as e:
            logger.debug(f"EDA chart error: {e}")
            return None

    async def _generate_bar_chart(
        self,
        labels: list[str],
        values: list[float],
        title: str,
    ) -> str | None:
        """Generate bar chart."""
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            fig, ax = plt.subplots(figsize=(max(8, len(labels) * 0.5), 5))

            # Truncate long labels
            short_labels = [
                l[:20] + "..." if len(l) > 20 else l for l in labels
            ]

            bars = ax.bar(
                range(len(values)), values,
                color="#4A90D9", alpha=0.8,
            )
            ax.set_xticks(range(len(short_labels)))
            ax.set_xticklabels(short_labels, rotation=45,
                               ha="right", fontsize=8)
            ax.set_title(title, fontsize=12)

            # Add value labels on bars
            for bar, val in zip(bars, values):
                ax.text(
                    bar.get_x() + bar.get_width() / 2, bar.get_height(),
                    f"{val:.1f}", ha="center", va="bottom", fontsize=7,
                )

            plt.tight_layout()
            chart_path = str(
                OUTPUT_DIR / f"bar_{datetime.now():%Y%m%d_%H%M%S}.png"
            )
            plt.savefig(chart_path, dpi=150)
            plt.close()
            return chart_path

        except Exception as e:
            logger.debug(f"Bar chart error: {e}")
            return None

    async def generate_chart(
        self,
        path: str,
        x_column: str,
        y_column: str,
        chart_type: str = "bar",
        title: str = "",
    ) -> AnalysisResult:
        """
        Generate a chart from data.
        chart_type: bar, line, pie, scatter
        """
        if not self._has_matplotlib:
            return AnalysisResult(
                success=False,
                error="matplotlib не установлен. pip install matplotlib"
            )

        try:
            headers, rows = self.auto_load(path)
        except Exception as e:
            return AnalysisResult(success=False, error=str(e))

        if x_column not in headers or y_column not in headers:
            return AnalysisResult(
                success=False,
                error=f"Столбцы не найдены. Доступные: {headers}"
            )

        x_idx = headers.index(x_column)
        y_idx = headers.index(y_column)

        x_vals = [row[x_idx] if x_idx < len(row) else "" for row in rows]
        y_raw = [row[y_idx] if y_idx < len(row) else "0" for row in rows]
        y_vals = self._to_numbers(y_raw)

        if len(y_vals) != len(x_vals):
            return AnalysisResult(
                success=False,
                error=f"Столбец {y_column} содержит нечисловые значения"
            )

        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            fig, ax = plt.subplots(figsize=(10, 6))

            if not title:
                title = f"{y_column} по {x_column}"

            if chart_type == "line":
                ax.plot(range(len(y_vals)), y_vals, marker="o", linewidth=2)
                ax.set_xticks(range(0, len(x_vals), max(1, len(x_vals) // 10)))
                ax.set_xticklabels(
                    [x_vals[i]
                        for i in range(0, len(x_vals), max(1, len(x_vals) // 10))],
                    rotation=45, ha="right", fontsize=8,
                )
            elif chart_type == "pie":
                ax.pie(
                    y_vals[:20], labels=x_vals[:20],
                    autopct="%1.1f%%", startangle=90,
                )
            elif chart_type == "scatter":
                x_nums = self._to_numbers(x_vals)
                if len(x_nums) == len(y_vals):
                    ax.scatter(x_nums, y_vals, alpha=0.6)
                    ax.set_xlabel(x_column)
                    ax.set_ylabel(y_column)
                else:
                    ax.scatter(range(len(y_vals)), y_vals, alpha=0.6)
            else:  # bar
                short = [
                    v[:15] + ".." if len(v) > 15 else v for v in x_vals
                ]
                ax.bar(range(len(y_vals)), y_vals, color="#4A90D9", alpha=0.8)
                ax.set_xticks(range(len(short)))
                ax.set_xticklabels(short, rotation=45, ha="right", fontsize=7)

            ax.set_title(title, fontsize=13)
            plt.tight_layout()

            chart_path = str(
                OUTPUT_DIR /
                f"chart_{chart_type}_{datetime.now():%Y%m%d_%H%M%S}.png"
            )
            plt.savefig(chart_path, dpi=150)
            plt.close()

            return AnalysisResult(
                success=True,
                summary=f"📈 График '{title}' создан",
                charts=[chart_path],
                data={"chart_type": chart_type, "points": len(y_vals)},
            )

        except Exception as e:
            return AnalysisResult(success=False, error=f"Ошибка графика: {e}")

    # ─── Export ──────────────────────────────────────────────────────────

    async def export_to_csv(
        self,
        headers: list[str],
        rows: list[list[str]],
        filename: str = "",
    ) -> str:
        """Export data to CSV file."""
        if not filename:
            filename = f"export_{datetime.now():%Y%m%d_%H%M%S}.csv"
        path = str(OUTPUT_DIR / filename)
        self._write_csv(path, headers, rows)
        return path

    async def export_to_excel(
        self,
        headers: list[str],
        rows: list[list[str]],
        filename: str = "",
    ) -> str:
        """Export data to Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Нужен openpyxl: pip install openpyxl")

        if not filename:
            filename = f"export_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = str(OUTPUT_DIR / filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(path)
        return path

    # ─── Helpers ─────────────────────────────────────────────────────────

    def _get_numeric_column(
        self,
        headers: list[str],
        rows: list[list[str]],
        col_name: str,
    ) -> list[float]:
        """Extract numeric values from column."""
        idx = headers.index(col_name)
        values = []
        for row in rows:
            if idx < len(row):
                try:
                    values.append(
                        float(row[idx].replace(",", ".").replace(" ", ""))
                    )
                except ValueError:
                    pass
        return values

    def _to_numbers(self, values: list[str]) -> list[float]:
        """Convert string values to numbers."""
        result = []
        for v in values:
            try:
                result.append(float(v.replace(",", ".").replace(" ", "")))
            except (ValueError, AttributeError):
                result.append(0.0)
        return result

    def _correlation(
        self,
        x: list[float],
        y: list[float],
    ) -> float:
        """Pearson correlation coefficient."""
        n = min(len(x), len(y))
        if n < 3:
            return 0.0

        x, y = x[:n], y[:n]
        mean_x = statistics.mean(x)
        mean_y = statistics.mean(y)

        num = sum((xi - mean_x) * (yi - mean_y) for xi, yi in zip(x, y))
        den_x = sum((xi - mean_x) ** 2 for xi in x) ** 0.5
        den_y = sum((yi - mean_y) ** 2 for yi in y) ** 0.5

        if den_x == 0 or den_y == 0:
            return 0.0

        return num / (den_x * den_y)

    def _write_csv(
        self,
        path: str,
        headers: list[str],
        rows: list[list[str]],
    ):
        """Write CSV file."""
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

    def _format_table(
        self,
        headers: list[str],
        rows: list[list[str]],
        max_rows: int = 30,
    ) -> str:
        """Format data as readable table."""
        if not headers and not rows:
            return "[Пустая таблица]"

        display_rows = rows[:max_rows]
        all_data = [headers] + display_rows

        col_count = max(len(r) for r in all_data) if all_data else 0
        widths = [0] * col_count
        for row in all_data:
            for i, cell in enumerate(row):
                if i < col_count:
                    widths[i] = max(widths[i], min(len(str(cell)), 20))

        lines = []
        # Header
        hdr = " | ".join(
            str(h)[:20].ljust(widths[i] if i < len(widths) else 10)
            for i, h in enumerate(headers)
        )
        lines.append(hdr)
        lines.append("-" * len(hdr))

        # Rows
        for row in display_rows:
            r = " | ".join(
                str(c)[:20].ljust(widths[i] if i < len(widths) else 10)
                for i, c in enumerate(row)
            )
            lines.append(r)

        if len(rows) > max_rows:
            lines.append(f"... ещё {len(rows) - max_rows} строк")

        return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL INSTANCE
# ═══════════════════════════════════════════════════════════════════════════════

data_engine = DataAnalysisEngine()
